from openpyxl import load_workbook
import subprocess
import time
import sys
import os



def AT_Command(runtimes):
    a = 0
    Script_path = sys.path[0]
    with open(Script_path+os.sep+'AT命令测试结果.txt','w') as f:
        f.write('')
    wb = load_workbook(Script_path + os.sep + '中移物联测试.xlsx')
    sheet = wb['at命令']
    max_rows = sheet.max_row
    for e in range(runtimes):
        for i in range(3,max_rows+1):
            at_command = sheet['D%d'%i].value
            except_result = str(sheet['E%d'%i].value)
            r =subprocess.Popen('adb shell send_at %s'%at_command,stdout = subprocess.PIPE,stderr=subprocess.PIPE,encoding='utf8')
            time.sleep(0.5)
            t = r.stdout.read()
            if except_result in t:
                print('pass')
            else:
                a = a + 1
                with open(Script_path+os.sep+'AT命令测试结果.txt','a') as f:
                    f.write(at_command+'执行失败   '+  '预期结果为：'+ except_result + '\n'+'                   实际结果为：'+ t.replace('\n','')+ '\n\n')
                print(at_command+'执行失败')
                break
        print('第 %d 轮测试结束'%(e+1))
        if a!=0:
            break
    with open(Script_path+os.sep+'AT命令测试结果.txt') as f:
        t = f.read()
        for i in range(3,max_rows+1):
            error_at_command =  sheet['D%d'%i].value
            pass_rate = 1 - (t.count(error_at_command))/runtimes
            sheet['G%d'%i].value = pass_rate
            sheet['F%d'%i].value = runtimes
    try:
        wb.save(Script_path + os.sep + '中移物联测试.xlsx')
    except:
        print('中移物联测试.xlsx文件未关闭')

# if __name__ =='__main__':
AT_Command(10000)